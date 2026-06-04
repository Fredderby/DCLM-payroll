from fastapi import FastAPI, Request, Depends, Form, File, UploadFile, HTTPException
from typing import List
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from app.api import auth, payroll
from app.core.config import settings
from app.services.cache_service import get_cache, set_cache
from app.core.database import get_db
from app.core.security import get_current_user, get_current_user_web
from app.models.user import User
import os
import urllib.parse

app = FastAPI(title="Payroll Processing System", version="1.0.0")

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# Templates
templates = Jinja2Templates(directory="templates")
templates.env.cache = None

import json

def tojson_filter(value):
    try:
        if hasattr(value, '__table__'):
            return json.dumps({c.name: getattr(value, c.name, None) for c in value.__table__.columns}, default=str)
        return json.dumps(value, default=str)
    except:
        return json.dumps(str(value))

# Custom Jinja filters
def format_month(value):
    """Convert '2024-01', '01-2024', '10-2026', or 'January 2024' to readable month name"""
    if not value:
        return value
    import datetime
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    try:
        val = str(value).strip()
        # Already formatted as 'Month YYYY'
        if any(m in val for m in month_names if m):
            return val
        parts = val.split('-')
        if len(parts) == 2:
            # Try both YYYY-MM and MM-YYYY
            p0 = parts[0].strip()
            p1 = parts[1].strip()
            if p0.isdigit() and len(p0) == 4:  # YYYY-MM
                year = p0
                month_num = p1.zfill(2)
            else:  # MM-YYYY
                month_num = p0.zfill(2)
                year = p1
            month_name = month_names[int(month_num)] if 1 <= int(month_num) <= 12 else month_num
            return f"{month_name} {year}"
        return val
    except:
        return value

templates.env.filters['format_month'] = format_month
templates.env.filters['tojson'] = tojson_filter

# Initialize database and create indexes on startup
from app.core.database import init_db
import asyncio

@app.on_event("startup")
async def startup_event():
    """Initialize application on startup."""
    # Run database initialization (tables + indexes) in executor to avoid blocking
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, init_db)

# API routers
app.include_router(auth.router, prefix="/auth", tags=["Authentication"])
app.include_router(payroll.router, prefix="/payroll", tags=["Payroll"])

@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return RedirectResponse(url="/login")

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    template = templates.get_template("login.html")
    rendered = template.render({"active_tab": "login"})
    return HTMLResponse(content=rendered, media_type="text/html")

@app.post("/login")
async def login(request: Request, email: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    from app.core.security import verify_password, create_access_token
    user = db.query(User).filter(User.email == email).first()
    if user and verify_password(password, user.hashed_password):
        # Create JWT token and set in cookie
        access_token = create_access_token(data={"sub": user.email})
        # Force fresh load after login to reinitialize all components
        response = RedirectResponse(url="/dashboard?_fresh=true", status_code=303)
        response.set_cookie(key="access_token", value=f"Bearer {access_token}", httponly=True)
        return response
    template = templates.get_template("login.html")
    rendered = template.render({"error": "Invalid credentials", "active_tab": "login"})
    return HTMLResponse(content=rendered, media_type="text/html")

@app.post("/register")
async def register(request: Request, email: str = Form(...), password: str = Form(...), confirm_password: str = Form(...), role: str = Form(...), first_name: str = Form(...), last_name: str = Form(...), terms: bool = Form(...), db: Session = Depends(get_db)):
    # Validate input
    if password != confirm_password:
        template = templates.get_template("login.html")
        rendered = template.render({"register_error": "Passwords do not match", "active_tab": "register"})
        return HTMLResponse(content=rendered, media_type="text/html")
    
    if len(password) < 8:
        template = templates.get_template("login.html")
        rendered = template.render({"register_error": "Password must be at least 8 characters long", "active_tab": "register"})
        return HTMLResponse(content=rendered, media_type="text/html")
    
    if not terms:
        template = templates.get_template("login.html")
        rendered = template.render({"register_error": "You must agree to the terms and conditions", "active_tab": "register"})
        return HTMLResponse(content=rendered, media_type="text/html")
    
    # Check if user already exists
    existing_user = db.query(User).filter(User.email == email).first()
    if existing_user:
        template = templates.get_template("login.html")
        rendered = template.render({"register_error": "Email already registered", "active_tab": "register"})
        return HTMLResponse(content=rendered, media_type="text/html")
    
    # Create user
    from app.core.security import get_password_hash
    hashed_password = get_password_hash(password)
    new_user = User(email=email, hashed_password=hashed_password, role=role)
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    # Success message
    template = templates.get_template("login.html")
    rendered = template.render({"register_success": "Account created successfully! Please login.", "active_tab": "login"})
    return HTMLResponse(content=rendered, media_type="text/html")

@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
        from app.models.employee import Employee
        from app.models.payroll import PayrollRecord
        from datetime import datetime
        from app.services.cache_service import get_cache, set_cache
        cache_key = "dash"
        cached = get_cache(cache_key)
        if cached:
            stats, recent_payslips = cached
        else:
            # Build stats for the dashboard - optimized queries
            from sqlalchemy import func as sa_func
            employee_count = db.query(sa_func.count(Employee.id)).scalar() or 0
            payslip_count = db.query(sa_func.count(PayrollRecord.id)).scalar() or 0
            total_payroll_sum = db.query(sa_func.coalesce(sa_func.sum(PayrollRecord.net_salary), 0)).scalar() or 0
            avg_net_salary = (total_payroll_sum / payslip_count) if payslip_count > 0 else 0
            
            # Distinct months - optimized
            months = db.query(PayrollRecord.month).distinct().all()
            months_active = len([m[0] for m in months if m[0]])
            
            stats = {
                "employee_count": employee_count,
                "payslip_count": payslip_count,
                "total_payroll": total_payroll_sum,
                "months_active": months_active,
                "upload_count": months_active,
                "avg_net_salary": round(avg_net_salary, 2),
                "email_sent": 0
            }
            
            # Recent payslips (last 5) - optimized with eager loading
            recent_payslips = db.query(PayrollRecord).order_by(PayrollRecord.id.desc()).limit(5).all()
            # Use a single query for all employees to avoid N+1
            employee_names = list(set(p.employee_name for p in recent_payslips))
            employees = {e.name: e for e in db.query(Employee).filter(Employee.name.in_(employee_names)).all()}
            for p in recent_payslips:
                p.employee = employees.get(p.employee_name)
            
            set_cache(cache_key, (stats, recent_payslips), ttl_seconds=120)
        
        template = templates.get_template("dashboard.html")
        rendered = template.render({
            "user": current_user,
            "stats": stats,
            "recent_payslips": recent_payslips,
            "activities": [],
            "now": datetime.now
        })
        return HTMLResponse(content=rendered, media_type="text/html")
    except HTTPException:
        template = templates.get_template("login.html")
        rendered = template.render({})
        return HTMLResponse(content=rendered, media_type="text/html")

@app.get("/upload", response_class=HTMLResponse)
async def upload_page(request: Request, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
        template = templates.get_template("upload.html")
        rendered = template.render({"user": current_user})
        return HTMLResponse(content=rendered, media_type="text/html")
    except HTTPException:
        template = templates.get_template("login.html")
        rendered = template.render({})
        return HTMLResponse(content=rendered, media_type="text/html")

def _build_excel_template(headers, sample_data, instructions, sheet_title, filename):
    """Helper to build and return an Excel template file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for col_idx, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx >= 2:
            cell.number_format = '#,##0.00'

    # Column widths
    for i in range(len(headers)):
        ws.column_dimensions[chr(65 + i)].width = max(18, len(headers[i]) + 4)

    for i, text in enumerate(instructions):
        cell = ws.cell(row=4 + i, column=1, value=text)
        cell.font = Font(bold=(i == 0), size=10, color="000000" if i == 0 else "666666")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return FileResponse(
        path=tmp.name,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.get("/upload/template/pastoral")
async def download_pastoral_template(request: Request):
    """Download the Pastoral Payroll Template.
    
    Contains Employee Name + standard pastoral earnings/deductions.
    Employee personal info syncs automatically from employee records.
    """
    try:
        headers = [
            "Employee Name",
            "Basic Salary",
            "Meals Monthly",
            "Responsibility Allowance",
            "COLA",
            "Leave Allowance",
            "SSNIT 5.5%",
            "PAYE",
            "10% Tithe",
            "Future Savings"
        ]
        sample_data = [
            "John Doe", 5000.00, 300.00, 200.00, 150.00, 0.00,
            0.00, 500.00, 500.00, 200.00
        ]
        instructions = [
            "INSTRUCTIONS - PASTORAL PAYROLL:",
            "1. Only Employee Name is required. Other employee info syncs from the employee database.",
            "2. Employee Name must exactly match the name in Employee Management (case-insensitive).",
            "3. All numeric fields are optional (leave blank or 0 if not applicable).",
            "4. Do not include currency symbols, commas, or special characters in numeric fields.",
            "5. The payroll month is selected separately in the upload form on the Upload Payroll page.",
            "6. Use this template for Pastoral staff only. For Non-Pastoral, download the other template."
        ]
        return _build_excel_template(headers, sample_data, instructions, "Pastoral Template", "pastoral_payroll_template.xlsx")
    except ImportError:
        return JSONResponse(content={"error": "openpyxl not available. Install with: pip install openpyxl"}, status_code=500)
    except Exception as e:
        return JSONResponse(content={"error": f"Failed to generate template: {str(e)}"}, status_code=500)


@app.get("/upload/template/non-pastoral")
async def download_non_pastoral_template(request: Request):
    """Download the Non-Pastoral Payroll Template.
    
    Contains Employee Name + non-pastoral earnings/deductions.
    Employee personal info syncs automatically from employee records.
    """
    try:
        headers = [
            "Employee Name",
            "Monthly Basic Salary",
            "Rent Monthly",
            "Utility Monthly",
            "Meals Monthly",
            "Transport Monthly",
            "COLA",
            "Leave Allowance",
            "SSNIT 5.5%",
            "PAYE",
            "10% Tithe",
            "Future Savings"
        ]
        sample_data = [
            "Jane Doe", 4000.00, 800.00, 400.00, 300.00, 500.00, 200.00,
            0.00, 220.00, 400.00, 400.00, 200.00
        ]
        instructions = [
            "INSTRUCTIONS - NON-PASTORAL PAYROLL:",
            "1. Only Employee Name is required. Other employee info syncs from the employee database.",
            "2. Employee Name must exactly match the name in Employee Management (case-insensitive).",
            "3. All numeric fields are optional (leave blank or 0 if not applicable).",
            "4. Do not include currency symbols, commas, or special characters in numeric fields.",
            "5. The payroll month is selected separately in the upload form on the Upload Payroll page.",
            "6. Use this template for Non-Pastoral staff only. For Pastoral, download the other template.",
            "7. SSNIT 5.5% = 5.5% of Monthly Basic Salary (social security contribution). Enter the exact value from your payroll spreadsheet."
        ]
        return _build_excel_template(headers, sample_data, instructions, "Non-Pastoral Template", "non_pastoral_payroll_template.xlsx")
    except ImportError:
        return JSONResponse(content={"error": "openpyxl not available. Install with: pip install openpyxl"}, status_code=500)
    except Exception as e:
        return JSONResponse(content={"error": f"Failed to generate template: {str(e)}"}, status_code=500)


@app.get("/upload/template/download")
async def download_payroll_template(request: Request):
    """Redirect to the pastoral template for backward compatibility."""
    # For backward compatibility: if someone clicks an old link, serve pastoral template
    return await download_pastoral_template(request)

@app.post("/upload")
async def upload_payroll(request: Request, file: UploadFile = File(...), month: str = Form(...), db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.services.excel_service import process_payroll_excel
    from app.services.payroll_service import create_payroll_record
    from app.models.employee import Employee
    from app.models.payroll import PayrollRecord
    from app.models.upload_history import UploadHistory
    from datetime import datetime, date
    import tempfile
    import os
    import time
    import logging
    
    logger = logging.getLogger(__name__)
    logger.info(f"UPLOAD START: filename={file.filename}, month={month}, user={current_user.email}")

    if not file.filename.endswith(('.xlsx', '.csv')):
        logger.warning(f"UPLOAD REJECTED: invalid file type {file.filename}")
        template = templates.get_template("upload.html")
        rendered = template.render({"user": current_user, "error": "Only .xlsx and .csv files allowed"})
        return HTMLResponse(content=rendered, media_type="text/html")

    original_filename = file.filename

    # Auto-detect month from filename if not explicitly provided
    # Try to extract month patterns like "January 2025", "Jan 2025", "2025-01" from filename
    if not month:
        import re
        month_patterns = [
            r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})',
            r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*\.?\s*(\d{4})',
            r'(\d{4})[-/](\d{2})',
            r'(\d{2})[-/](\d{4})',
        ]
        for pattern in month_patterns:
            match = re.search(pattern, original_filename, re.IGNORECASE)
            if match:
                groups = match.groups()
                if len(groups) == 2:
                    # Check if first group is a month name
                    month_names = {'january': 'January', 'february': 'February', 'march': 'March',
                                   'april': 'April', 'may': 'May', 'june': 'June', 'july': 'July',
                                   'august': 'August', 'september': 'September', 'october': 'October',
                                   'november': 'November', 'december': 'December',
                                   'jan': 'January', 'feb': 'February', 'mar': 'March',
                                   'apr': 'April', 'jun': 'June', 'jul': 'July', 'aug': 'August',
                                   'sep': 'September', 'oct': 'October', 'nov': 'November', 'dec': 'December'}
                    first = groups[0].lower()
                    if first in month_names:
                        month = f"{month_names[first]} {groups[1]}"
                        break
                    # Check if first is a year
                    try:
                        yr = int(groups[0])
                        if yr > 2000 and yr < 2100:
                            month_num = int(groups[1])
                            if 1 <= month_num <= 12:
                                from datetime import datetime as dt
                                month = dt(yr, month_num, 1).strftime("%B %Y")
                                break
                    except ValueError:
                        pass
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    max_retries = 3
    retry_delay = 2
    
    for attempt in range(max_retries):
        try:
            db.close()
            from app.core.database import SessionLocal
            db = SessionLocal()
            
            records = process_payroll_excel(tmp_path, month, filename=original_filename)
            processed = 0
            errors = []
            unmatched_records = []  # Names that didn't match any registered employee
            
            # Track IDs of employees that appear in the file
            matched_emp_ids = set()
            
            logger.info(f"UPLOAD PROCESSING: {len(records)} records extracted from file '{original_filename}' for month '{month}' by user '{current_user.email}'")
            
            for record in records:
                skip_reason = None
                try:
                    emp_no = str(record.get('employee_number', '') or '').strip()
                    emp_email = str(record.get('email', '') or '').strip()
                    emp_name = str(record.get('employee_name', '') or '').strip()
                    # Normalize: collapse multiple spaces, remove leading/trailing spaces
                    emp_name_norm = ' '.join(emp_name.split()).upper()
                    
                    if not emp_name_norm:
                        skip_reason = "Missing employee name"
                        unmatched_records.append({
                            'name': emp_no or emp_email or 'Unknown',
                            'number': emp_no,
                            'email': emp_email,
                            'reason': skip_reason
                        })
                        print(f"SKIP: {skip_reason} - emp_no={repr(emp_no)}")
                        continue
                    
                    employee = None
                    # Build lookup dict for faster matching
                    all_emps = db.query(Employee).all()
                    name_lookup = {}
                    for e in all_emps:
                        norm = ' '.join((e.name or '').split()).upper()
                        name_lookup[norm] = e
                    
                    # PRIMARY: Match by employee name (case-insensitive, normalized)
                    if emp_name_norm in name_lookup:
                        employee = name_lookup[emp_name_norm]
                    if not employee and emp_name_norm:
                        # Try partial match
                        employee = db.query(Employee).filter(
                            Employee.name.ilike(f'%{emp_name}%')
                        ).first()
                    if not employee and emp_name_norm:
                        # Try matching by individual name parts
                        name_parts = emp_name_norm.split()
                        for part in name_parts:
                            if len(part) > 2:
                                for norm, e in name_lookup.items():
                                    if part in norm:
                                        employee = e
                                        break
                                if employee:
                                    break
                    # FALLBACK: Match by employee number if name didn't match
                    if not employee and emp_no:
                        employee = db.query(Employee).filter(Employee.employee_number == emp_no).first()
                    # FALLBACK: Match by email if still no match
                    if not employee and emp_email:
                        employee = db.query(Employee).filter(Employee.email == emp_email).first()
                    
                    if not employee:
                        # No match found - skip and list for correction
                        skip_reason = f"No employee found matching name '{emp_name}'"
                        unmatched_records.append({
                            'name': emp_name or emp_no or emp_email or 'Unknown',
                            'number': emp_no,
                            'email': emp_email,
                            'reason': skip_reason
                        })
                        print(f"SKIP: {skip_reason} - emp_no={repr(emp_no)} emp_email={repr(emp_email)}")
                        continue
                    
                    print(f"MATCH OK: emp_name={repr(emp_name)} -> {employee.name} (ID={employee.id})")
                    
                    # Existing employee matched - update details from file if provided
                    if record.get('employee_name'): 
                        employee.name = str(record['employee_name']).strip()
                    if record.get('function'): 
                        employee.function = str(record['function']).strip()
                    if record.get('designation'): 
                        employee.designation = str(record['designation']).strip()
                    if record.get('location'): 
                        employee.location = str(record['location']).strip()
                    if record.get('bank_account'):
                        from app.services.pdf_service import parse_bank_account
                        bn, bk, bb = parse_bank_account(record.get('bank_account', ''))
                        if bn: employee.bank_number = bn
                        if bk: employee.bank_name = bk
                        if bb: employee.bank_branch = bb
                    db.commit()
                    
                    matched_emp_ids.add(employee.id)
                    
                    # Determine staff_category from record
                    staff_category = record.get('staff_category', 'pastoral')
                    
                    # Check if payroll record already exists for this employee + month
                    existing_payroll = db.query(PayrollRecord).filter(
                        PayrollRecord.employee_name == employee.name,
                        PayrollRecord.month == month
                    ).first()
                    
                    if existing_payroll:
                        # Update existing record instead of creating duplicate
                        existing_payroll.staff_category = staff_category
                        existing_payroll.basic_salary = record.get('basic_salary', existing_payroll.basic_salary)
                        existing_payroll.meals_monthly = record.get('meals_monthly', existing_payroll.meals_monthly)
                        existing_payroll.responsibility_allowance = record.get('responsibility_allowance', existing_payroll.responsibility_allowance)
                        existing_payroll.cola = record.get('cola', existing_payroll.cola)
                        existing_payroll.leave_allowance = record.get('leave_allowance', existing_payroll.leave_allowance)
                        existing_payroll.other_earnings = record.get('other_earnings', existing_payroll.other_earnings)
                        # Non-Pastoral earnings
                        existing_payroll.rent_monthly = record.get('rent_monthly', existing_payroll.rent_monthly)
                        existing_payroll.utility_monthly = record.get('utility_monthly', existing_payroll.utility_monthly)
                        existing_payroll.transport_monthly = record.get('transport_monthly', existing_payroll.transport_monthly)
                        existing_payroll.paye = record.get('paye', existing_payroll.paye)
                        existing_payroll.tithe = record.get('tithe', existing_payroll.tithe)
                        existing_payroll.future_savings = record.get('future_savings', existing_payroll.future_savings)
                        existing_payroll.other_deductions = record.get('other_deductions', existing_payroll.other_deductions)
                        existing_payroll.employer_contribution = record.get('employer_contribution', existing_payroll.employer_contribution)
                        existing_payroll.employee_pf = record.get('employee_pf', existing_payroll.employee_pf)
                        existing_payroll.ssnit_deduction = record.get('ssnit_deduction', existing_payroll.ssnit_deduction)
                        from app.services.payroll_service import calculate_payroll_totals
                        earnings = {
                            'basic_salary': existing_payroll.basic_salary,
                            'meals_monthly': existing_payroll.meals_monthly,
                            'responsibility_allowance': existing_payroll.responsibility_allowance,
                            'cola': existing_payroll.cola,
                            'leave_allowance': existing_payroll.leave_allowance,
                            'other_earnings': existing_payroll.other_earnings,
                            'rent_monthly': existing_payroll.rent_monthly,
                            'utility_monthly': existing_payroll.utility_monthly,
                            'transport_monthly': existing_payroll.transport_monthly,
                        }
                        deductions = {
                            'paye': existing_payroll.paye,
                            'tithe': existing_payroll.tithe,
                            'future_savings': existing_payroll.future_savings,
                            'other_deductions': existing_payroll.other_deductions,
                            'employee_pf': existing_payroll.employee_pf,
                            'ssnit_deduction': existing_payroll.ssnit_deduction,
                        }
                        total_earnings, total_deductions, net_salary = calculate_payroll_totals(earnings, deductions)
                        existing_payroll.total_earnings = total_earnings
                        existing_payroll.total_deductions = total_deductions
                        existing_payroll.net_salary = net_salary
                        # Clear old PDF so it gets regenerated on demand
                        if existing_payroll.pdf_generated:
                            try:
                                if os.path.exists(existing_payroll.pdf_generated):
                                    os.remove(existing_payroll.pdf_generated)
                            except:
                                pass
                            existing_payroll.pdf_generated = None
                        db.commit()
                    else:
                        # Create payroll record (no PDF generation during upload)
                        # Map record fields to employee object for create_payroll_record
                        employee.name = record.get('employee_name', employee.name)
                        employee.email = record.get('email', employee.email)
                        employee.function = record.get('function', employee.function)
                        employee.designation = record.get('designation', employee.designation)
                        employee.location = record.get('location', employee.location)
                        employee.bank_account = record.get('bank_account', getattr(employee, 'bank_account', ''))
                        employee.ssnit_number = record.get('ssnit_number', getattr(employee, 'ssnit_number', ''))
                        employee.date_joined = record.get('date_joined', getattr(employee, 'date_joined', None))
                        payroll_record = create_payroll_record(db, employee, record)
                        db.commit()
                    
                    processed += 1
                except Exception as e:
                    db.rollback()
                    errors.append(f"Error processing {record.get('employee_name', 'Unknown')}: {str(e)[:80]}")
            
            # Identify registered employees NOT in this upload (possibly terminated)
            all_emp_ids = {r.id for r in db.query(Employee.id).all()}
            missing_emp_ids = all_emp_ids - matched_emp_ids
            missing_employees = []
            if missing_emp_ids:
                missing_emps = db.query(Employee).filter(Employee.id.in_(missing_emp_ids)).order_by(Employee.name).all()
                missing_employees = [e for e in missing_emps if e.name]
            
            # Invalidate dashboard cache
            from app.services.cache_service import clear_cache
            clear_cache()
            
            # Clear dashboard cache so stats update immediately
            from app.services.cache_service import clear_cache
            clear_cache()
            
            # Record upload history with detailed tracking
            try:
                import json
                skip_details = []
                for ur in unmatched_records:
                    skip_details.append({
                        'name': ur.get('name', 'Unknown'),
                        'employee_number': ur.get('number', ''),
                        'email': ur.get('email', ''),
                        'reason': ur.get('reason', 'No matching employee found')
                    })
                for mer in missing_employees:
                    skip_details.append({
                        'name': mer.name,
                        'employee_number': mer.employee_number,
                        'email': mer.email,
                        'reason': 'Registered employee not found in upload file (may be terminated)'
                    })
                
                total_records_in_file = len(records) if records else 0
                upload_history = UploadHistory(
                    file_name=original_filename,
                    uploaded_by=current_user.id,
                    month=month,
                    total_employees=total_records_in_file,
                    imported_count=processed,
                    skipped_count=len(unmatched_records) + len(missing_employees),
                    skip_reasons=json.dumps(skip_details),
                    status="success" if processed > 0 and len(errors) == 0 else "partial_failure" if processed > 0 else "failed"
                )
                db.add(upload_history)
                db.commit()
            except Exception as log_err:
                print(f"Warning: Could not log upload history: {log_err}")
                db.rollback()
            
            # Build summary
            summary_parts = []
            if processed:
                summary_parts.append(f"Successfully uploaded {processed} payroll record(s) for {month}")
            if unmatched_records:
                summary_parts.append(f"{len(unmatched_records)} name(s) did not match any registered employee (skipped)")
            if missing_employees:
                summary_parts.append(f"{len(missing_employees)} registered employee(s) not in this file (may be terminated)")
            
            msg = ". ".join(summary_parts) + "." if summary_parts else "No data processed."

            template = templates.get_template("upload.html")
            rendered = template.render({
                "user": current_user,
                "message": msg,
                "errors": errors[:5],
                "unmatched_records": unmatched_records,
                "missing_employees": missing_employees,
                "processed_count": processed
            })
            return HTMLResponse(content=rendered, media_type="text/html")
            
        except ValueError as e:
            db.rollback()
            template = templates.get_template("upload.html")
            rendered = template.render({"user": current_user, "error": str(e)})
            return HTMLResponse(content=rendered, media_type="text/html")
        except Exception as e:
            db.rollback()
            template = templates.get_template("upload.html")
            rendered = template.render({"user": current_user, "error": f"Upload failed due to an unexpected error. Please verify your file format and try again."})
            return HTMLResponse(content=rendered, media_type="text/html")
        finally:
            try:
                os.unlink(tmp_path)
            except:
                pass
    
    template = templates.get_template("upload.html")
    rendered = template.render({"user": current_user, "error": "Upload failed after multiple retry attempts"})
    return HTMLResponse(content=rendered, media_type="text/html")

@app.get("/staff", response_class=HTMLResponse)
async def staff_page(request: Request, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
        from app.models.employee import Employee
        
        # Get all staff members
        employees = db.query(Employee).all()
        
        template = templates.get_template("staff.html")
        rendered = template.render({"user": current_user, "employees": employees})
        return HTMLResponse(content=rendered, media_type="text/html")
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    except Exception as e:
        template = templates.get_template("staff.html")
        rendered = template.render({"user": get_current_user_web(request, db), "employees": [], "error": f"Error: {str(e)}"})
        return HTMLResponse(content=rendered, media_type="text/html")

@app.get("/staff/template/download")
async def download_employee_template(request: Request, db: Session = Depends(get_db)):
    """Download a CSV template for employee bulk upload - rebuilt for reliability."""
    try:
        get_current_user_web(request, db)
    except Exception:
        return JSONResponse(status_code=401, content={"error": "Not authenticated"})
    import csv
    import io
    
    fieldnames = [
        "employee_number", "name", "email", "function", "designation",
        "location", "ssnit_number", "tax_relief", "employer_contribution",
        "bank_number", "bank_name", "bank_branch", "date_joined"
    ]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerow({
        "employee_number": "EMP001",
        "name": "John Doe",
        "email": "john@example.com",
        "function": "Finance",
        "designation": "Accountant",
        "location": "Accra",
        "ssnit_number": "SSNIT123456",
        "tax_relief": "500",
        "employer_contribution": "0",
        "bank_number": "1234567890",
        "bank_name": "GCB Bank",
        "bank_branch": "Head Office",
        "date_joined": "2024-01-15"
    })
    csv_bytes = output.getvalue().encode('utf-8-sig')
    
    return Response(
        content=csv_bytes,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=employee_template.csv"}
    )

@app.post("/staff/upload")
async def upload_employees(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Upload CSV or Excel file to bulk add employees. Duplicates are automatically skipped."""
    try:
        current_user = get_current_user_web(request, db)
    except Exception:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.employee import Employee
    import pandas as pd
    from io import BytesIO
    from datetime import datetime
    
    if not file.filename:
        employees = db.query(Employee).all()
        template = templates.get_template("staff.html")
        return HTMLResponse(content=template.render({"user": current_user, "employees": employees, "error": "No file provided"}), media_type="text/html")
    
    try:
        contents = await file.read()
        ext = file.filename.rsplit('.', 1)[-1].lower()
        
        if ext == 'csv':
            df = pd.read_csv(BytesIO(contents), dtype=str)
        elif ext in ('xlsx', 'xls'):
            df = pd.read_excel(BytesIO(contents), dtype=str)
        else:
            employees = db.query(Employee).all()
            template = templates.get_template("staff.html")
            return HTMLResponse(content=template.render({"user": current_user, "employees": employees, "error": "Unsupported file format. Please upload CSV or Excel files."}), media_type="text/html")
        
        # Normalize column names
        df.columns = [c.strip().lower().replace(' ', '_').replace('-', '_') for c in df.columns]
        
        # Map expected columns
        col_map = {
            'employee_number': ['employee_number', 'employee_no', 'emp_number', 'empid', 'employee_id'],
            'name': ['name', 'employee_name', 'full_name', 'employeename'],
            'email': ['email', 'e_mail', 'email_address'],
            'function': ['function', 'department', 'dept', 'business_unit'],
            'designation': ['designation', 'title', 'job_title', 'position'],
            'location': ['location', 'office', 'branch', 'work_location'],
            'ssnit_number': ['ssnit_number', 'ssnit', 'social_security', 'ssn'],
            'tax_relief': ['tax_relief', 'taxrelief', 'tax_relief_amount'],
            'employer_contribution': ['employer_contribution', 'employer_contr', 'employercontribution'],
            'bank_number': ['bank_number', 'bank_no', 'account_number', 'bankaccount', 'bank_acc'],
            'bank_name': ['bank_name', 'bank', 'bankname'],
            'bank_branch': ['bank_branch', 'branch', 'bankbranch'],
            'date_joined': ['date_joined', 'joined_date', 'start_date', 'employment_date', 'doj']
        }
        
        normalized_cols = {}
        for std_name, aliases in col_map.items():
            for col in df.columns:
                if col in aliases:
                    normalized_cols[col] = std_name
                    break
        
        df.rename(columns=normalized_cols, inplace=True)
        
        standard_cols = list(col_map.keys())
        for col in df.columns:
            if col not in standard_cols:
                df.drop(columns=[col], inplace=True)
        
        if df.empty or 'name' not in df.columns or 'employee_number' not in df.columns:
            employees = db.query(Employee).all()
            template = templates.get_template("staff.html")
            return HTMLResponse(content=template.render({"user": current_user, "employees": employees, "error": "File must contain at least 'Name' and 'Employee Number' columns."}), media_type="text/html")
        
        # Build lookup sets from existing employees for multi-column dedup
        existing_emps = db.query(Employee).all()
        key_cols = ['employee_number', 'name', 'email', 'bank_number']
        existing_lookups = {}
        for col in key_cols:
            existing_lookups[col] = set()
        for emp in existing_emps:
            for col in key_cols:
                val = getattr(emp, col, None)
                if val:
                    existing_lookups[col].add(str(val).strip().lower())
        
        added_count = 0
        skipped_count = 0
        
        for _, row in df.iterrows():
            emp_num = str(row.get('employee_number', '')).strip()
            name = str(row.get('name', '')).strip()
            email = str(row.get('email', '')).strip()
            bank_num = str(row.get('bank_number', '')).strip()
            
            if not emp_num or not name:
                skipped_count += 1
                continue
            
            match_count = 0
            if emp_num.lower() in existing_lookups['employee_number']: match_count += 1
            if name.lower() in existing_lookups['name']: match_count += 1
            if email.lower() and email.lower() in existing_lookups['email']: match_count += 1
            if bank_num.lower() and bank_num.lower() in existing_lookups['bank_number']: match_count += 1
            
            if match_count >= 3:
                skipped_count += 1
                continue
            
            date_obj = None
            date_str = str(row.get('date_joined', '')).strip()
            if date_str and date_str.lower() not in ('nan', 'none', ''):
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                    try:
                        date_obj = datetime.strptime(date_str, fmt).date()
                        break
                    except ValueError:
                        pass
            
            def parse_float(val, default=0):
                try:
                    v = str(val).strip()
                    if v and v.lower() not in ('nan', 'none', ''):
                        return float(v.replace(',', ''))
                except (ValueError, TypeError):
                    pass
                return default
            
            employee = Employee(
                employee_number=emp_num,
                name=name,
                email=str(row.get('email', '')).strip(),
                function=str(row.get('function', '')).strip(),
                designation=str(row.get('designation', '')).strip(),
                location=str(row.get('location', '')).strip(),
                ssnit_number=str(row.get('ssnit_number', '')).strip(),
                tax_relief=str(row.get('tax_relief', '')).strip(),
                employer_contribution=parse_float(row.get('employer_contribution')),
                bank_number=str(row.get('bank_number', '')).strip(),
                bank_name=str(row.get('bank_name', '')).strip(),
                bank_branch=str(row.get('bank_branch', '')).strip(),
                date_joined=date_obj
            )
            db.add(employee)
            existing_lookups['employee_number'].add(emp_num.lower())
            added_count += 1
        
        db.commit()
        
        employees = db.query(Employee).all()
        template = templates.get_template("staff.html")
        msg = f"Successfully added {added_count} employee(s)."
        if skipped_count > 0:
            msg += f" {skipped_count} row(s) skipped (duplicates or invalid data)."
        return HTMLResponse(content=template.render({"user": current_user, "employees": employees, "success": msg}), media_type="text/html")
    
    except Exception as e:
        db.rollback()
        employees = db.query(Employee).all()
        template = templates.get_template("staff.html")
        return HTMLResponse(content=template.render({"user": current_user, "employees": employees, "error": f"Upload failed: {str(e)}"}), media_type="text/html")

@app.post("/staff/add")
@app.post("/staff/create")
async def add_staff(request: Request, employee_number: str = Form(...), name: str = Form(...), 
                   email: str = Form(...), function: str = Form(default=""),
                   designation: str = Form(default=""), location: str = Form(default=""),
                   ssnit_number: str = Form(default=""), tax_relief: str = Form(default=""),
                   employer_contribution: float = Form(default=0), bank_number: str = Form(default=""),
                   bank_name: str = Form(default=""), bank_branch: str = Form(default=""),
                   date_joined: str = Form(default=None), db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.employee import Employee
    from datetime import datetime
    
    # Check if employee exists
    existing = db.query(Employee).filter(
    (Employee.email == email) | (Employee.employee_number == employee_number)
    ).first()
    
    if existing:
        employees = db.query(Employee).all()
        template = templates.get_template("staff.html")
        rendered = template.render({"user": current_user, "employees": employees, "error": "Employee already exists"})
        return HTMLResponse(content=rendered, media_type="text/html")
    
    try:
        date_obj = None
        if date_joined:
            date_obj = datetime.strptime(date_joined, "%Y-%m-%d").date()
        
        employee = Employee(
            employee_number=employee_number,
            name=name,
            email=email,
            function=function,
            designation=designation,
            location=location,
            ssnit_number=ssnit_number,
            tax_relief=tax_relief,
            employer_contribution=employer_contribution or 0,
            bank_number=bank_number,
            bank_name=bank_name,
            bank_branch=bank_branch,
            date_joined=date_obj
        )
        db.add(employee)
        db.commit()
        
        employees = db.query(Employee).all()
        template = templates.get_template("staff.html")
        rendered = template.render({"user": current_user, "employees": employees, "success": f"Staff member {name} added successfully"})
        return HTMLResponse(content=rendered, media_type="text/html")
    except Exception as e:
        db.rollback()
        employees = db.query(Employee).all()
        template = templates.get_template("staff.html")
        rendered = template.render({"user": current_user, "employees": employees, "error": f"Failed to add staff: {str(e)}"})
        return HTMLResponse(content=rendered, media_type="text/html")

@app.get("/staff/{staff_id}/details")
async def staff_details(request: Request, staff_id: int, db: Session = Depends(get_db)):
    """Return employee details as JSON for view/edit modals"""
    try:
        current_user = get_current_user_web(request, db)
    except Exception:
        return JSONResponse(status_code=401, content={"error": "Not authenticated"})
    
    from app.models.employee import Employee
    from datetime import date as dt_date
    
    employee = db.query(Employee).filter(Employee.id == staff_id).first()
    if not employee:
        return JSONResponse(status_code=404, content={"error": "Staff member not found"})
    
    djoin = ""
    if employee.date_joined:
        if isinstance(employee.date_joined, (dt_date,)):
            djoin = employee.date_joined.isoformat()
        else:
            djoin = str(employee.date_joined)
    
    return {
        "id": employee.id,
        "name": employee.name,
        "employee_number": employee.employee_number,
        "email": employee.email or "",
        "function": employee.function or "",
        "designation": employee.designation or "",
        "location": employee.location or "",
        "ssnit_number": employee.ssnit_number or "",
        "tax_relief": employee.tax_relief or "",
        "employer_contribution": float(employee.employer_contribution or 0),
        "bank_number": employee.bank_number or "",
        "bank_name": employee.bank_name or "",
        "bank_branch": employee.bank_branch or "",
        "date_joined": djoin
    }

@app.post("/staff/update/{staff_id}")
async def update_staff(request: Request, staff_id: int, employee_number: str = Form(...), name: str = Form(...), 
                       email: str = Form(...), function: str = Form(default=""),
                       designation: str = Form(default=""), location: str = Form(default=""),
                       ssnit_number: str = Form(default=""), tax_relief: str = Form(default=""),
                       employer_contribution: float = Form(default=0), bank_number: str = Form(default=""),
                       bank_name: str = Form(default=""), bank_branch: str = Form(default=""),
                       date_joined: str = Form(default=None), db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.employee import Employee
    from datetime import datetime
    
    try:
        staff = db.query(Employee).filter(Employee.id == staff_id).first()
        if not staff:
            raise HTTPException(status_code=404, detail="Staff member not found")
        
        # Check if new email/employee_number already exists for another employee
        existing = db.query(Employee).filter(
            (Employee.id != staff_id) &
            ((Employee.email == email) | (Employee.employee_number == employee_number))
        ).first()
        
        if existing:
            employees = db.query(Employee).all()
            template = templates.get_template("staff.html")
            rendered = template.render({"user": current_user, "staff": staff, "employees": employees, 
                                      "error": "Email or employee number already exists"})
            return HTMLResponse(content=rendered, media_type="text/html")
        
        # Update employee
        staff.employee_number = employee_number
        staff.name = name
        staff.email = email
        staff.function = function
        staff.designation = designation
        staff.location = location
        staff.ssnit_number = ssnit_number
        staff.tax_relief = tax_relief
        staff.employer_contribution = employer_contribution or 0
        staff.bank_number = bank_number
        staff.bank_name = bank_name
        staff.bank_branch = bank_branch
        
        if date_joined:
            staff.date_joined = datetime.strptime(date_joined, "%Y-%m-%d").date()
        
        db.commit()
        
        employees = db.query(Employee).all()
        template = templates.get_template("staff.html")
        rendered = template.render({"user": current_user, "employees": employees, "success": f"Staff member {name} updated successfully"})
        return HTMLResponse(content=rendered, media_type="text/html")
    except Exception as e:
        db.rollback()
        employees = db.query(Employee).all()
        staff = db.query(Employee).filter(Employee.id == staff_id).first()
        template = templates.get_template("staff.html")
        rendered = template.render({"user": current_user, "staff": staff, "employees": employees, "error": f"Failed to update staff: {str(e)}"})
        return HTMLResponse(content=rendered, media_type="text/html")

@app.post("/staff/delete/{staff_id}")
async def delete_staff(request: Request, staff_id: int, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.employee import Employee
    
    try:
        staff = db.query(Employee).filter(Employee.id == staff_id).first()
        if not staff:
            employees = db.query(Employee).all()
            template = templates.get_template("staff.html")
            rendered = template.render({"user": current_user, "employees": employees, "error": "Staff member not found"})
            return HTMLResponse(content=rendered, media_type="text/html")
        
        staff_name = staff.name
        db.delete(staff)
        db.commit()
        
        employees = db.query(Employee).all()
        template = templates.get_template("staff.html")
        rendered = template.render({"user": current_user, "employees": employees, "success": f"Staff member {staff_name} deleted successfully"})
        return HTMLResponse(content=rendered, media_type="text/html")
    except Exception as e:
        db.rollback()
        employees = db.query(Employee).all()
        template = templates.get_template("staff.html")
        rendered = template.render({"user": current_user, "employees": employees, "error": f"Failed to delete staff: {str(e)}"})
        return HTMLResponse(content=rendered, media_type="text/html")

@app.get("/payslips", response_class=HTMLResponse)
async def payslips_page(request: Request, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
        from app.models.payroll import PayrollRecord
        from app.models.employee import Employee
        from app.services.cache_service import get_cache, set_cache
        from sqlalchemy import func as sa_func
        
        # Get distinct months for filter (cached separately)
        month_cache_key = "payslip_months"
        months = get_cache(month_cache_key)
        if not months:
            payroll_months = db.query(PayrollRecord.month).distinct().order_by(PayrollRecord.month.desc()).all()
            months = [m[0] for m in payroll_months if m[0]]
            set_cache(month_cache_key, months, ttl_seconds=300)
        
        selected_month = request.query_params.get("month", "")
        
        # Query all payslips (no pagination limit to ensure all are visible)
        query = db.query(PayrollRecord)
        if selected_month:
            query = query.filter(PayrollRecord.month == selected_month)
        
        total_payslips = query.count()
        payslips = query.order_by(PayrollRecord.employee_name).all()
        
        # Batch-load employee info (N+1 fix) with robust matching
        all_employees = {e.name: e for e in db.query(Employee).all()}
        for p in payslips:
            emp_name = (p.employee_name or '').strip()
            emp = all_employees.get(emp_name)
            if not emp:
                # Try case-insensitive match
                emp_name_upper = emp_name.upper()
                for e_name, e_obj in all_employees.items():
                    if (e_name or '').upper() == emp_name_upper:
                        emp = e_obj
                        break
            if not emp:
                # Try normalized whitespace match
                emp_name_norm = ' '.join(emp_name.split())
                for e_name, e_obj in all_employees.items():
                    if ' '.join((e_name or '').split()) == emp_name_norm:
                        emp = e_obj
                        break
            p.employee = emp
        
        # Generate stats from all records
        if selected_month:
            net_total = db.query(sa_func.coalesce(sa_func.sum(PayrollRecord.net_salary), 0)).filter(PayrollRecord.month == selected_month).scalar() or 0
        else:
            net_total = db.query(sa_func.coalesce(sa_func.sum(PayrollRecord.net_salary), 0)).scalar() or 0
        total_net_salary = float(net_total)
        total_earnings = sum(p.total_earnings or 0 for p in payslips)
        total_deductions = sum(p.total_deductions or 0 for p in payslips)
        
        template = templates.get_template("payslips.html")
        rendered = template.render({
            "user": current_user, 
            "payslips": payslips,
            "months": months,
            "selected_month": selected_month,
            "total_payslips": total_payslips,
            "total_net_salary": round(total_net_salary, 2),
            "total_earnings": round(total_earnings, 2),
            "total_deductions": round(total_deductions, 2),
            "success": request.query_params.get("success"),
            "error": request.query_params.get("error")
        })
        return HTMLResponse(content=rendered, media_type="text/html")
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    except Exception as e:
        template = templates.get_template("payslips.html")
        rendered = template.render({"error": f"Error loading payslips: {str(e)}", "payslips": [], "months": [], "selected_month": ""})
        return HTMLResponse(content=rendered, media_type="text/html")


@app.get("/payslips/search")
async def payslips_search(request: Request, db: Session = Depends(get_db), q: str = "", month: str = "", page: int = 1):
    """Search payslips with pagination for AJAX calls"""
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return JSONResponse(content={"error": "Not authenticated"}, status_code=401)
    
    from app.models.payroll import PayrollRecord
    from sqlalchemy import func as sa_func
    
    per_page = 50
    query = db.query(PayrollRecord)
    
    if q:
        query = query.filter(PayrollRecord.employee_name.like(f"%{q}%"))
    if month:
        query = query.filter(PayrollRecord.month == month)
    
    total = query.count()
    records = query.order_by(PayrollRecord.employee_name).offset((page - 1) * per_page).limit(per_page).all()
    
    return JSONResponse(content={
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page),
        "records": [{"id": r.id, "employee_name": r.employee_name, "month": r.month, 
                      "net_salary": r.net_salary, "pdf_generated": r.pdf_generated} for r in records]
    })

@app.get("/payslips/{payroll_id}/data")
async def payslip_data(payroll_id: int, request: Request, db: Session = Depends(get_db)):
    """Return payslip data as JSON for preview modal"""
    try:
        current_user = get_current_user_web(request, db)
    except Exception:
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=401, content={"error": "Not authenticated"})
    
    from app.models.payroll import PayrollRecord
    from app.models.employee import Employee
    from app.models.loan import Loan
    from app.services.pdf_service import generate_payslip_pdf
    
    payroll = db.query(PayrollRecord).filter(PayrollRecord.id == payroll_id).first()
    if not payroll:
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=404, content={"error": "Payslip not found"})
    
    # Generate PDF on demand if not already generated (for preview button to work)
    if not payroll.pdf_generated or not isinstance(payroll.pdf_generated, str) or not os.path.exists(str(payroll.pdf_generated or "")):
        try:
            pdf_path = generate_payslip_pdf(db, payroll.id)
            if pdf_path:
                payroll.pdf_generated = pdf_path
                db.commit()
        except Exception as e:
            pass  # PDF generation failed, preview still works
    
    employee = db.query(Employee).filter(Employee.name == payroll.employee_name).first()
    
    emp_data = {
        "name": employee.name if employee else payroll.employee_name or "Unknown",
        "employee_number": employee.employee_number if employee else "N/A",
        "designation": employee.designation if employee else "",
        "function": employee.function if employee else "",
        "location": employee.location if employee else "",
        "bank_name": employee.bank_name if employee else "N/A",
        "bank_branch": employee.bank_branch if employee else "",
        "bank_number": employee.bank_number if employee else "N/A",
        "email": employee.email if employee else "",
        "date_joined": str(employee.date_joined) if employee and employee.date_joined else "",
        "ssnit_number": employee.ssnit_number if employee else "",
        }

    # Fetch loans for this employee (matched by name)
    loans_data = []
    if employee and employee.name:
        loan_records = db.query(Loan).filter(Loan.employee_name == employee.name).filter(Loan.status != "Completed").all()
        for loan in loan_records:
            months_remaining = max(0, (loan.months_to_pay or 1) - (loan.months_paid or 0))
            loans_data.append({
                "bank_name": loan.bank_name or "",
                "loan_amount": float(loan.loan_amount or 0),
                "amount_paid": float(loan.amount_paid or 0),
                "months_remaining": months_remaining
            })
    
    # Net salary in words
    net_words = ""
    if payroll.net_salary:
        try:
            from app.services.payroll_service import number_to_words
            net_words = number_to_words(payroll.net_salary)
        except Exception:
            net_words = ""

    # Determine staff category for component segregation
    staff_category = payroll.staff_category or "pastoral"
    
    return {
        "id": payroll.id,
        "month": payroll.month,
        "employee_name": payroll.employee_name or "N/A",
        "employee_number": emp_data.get("employee_number", "N/A"),
        "email": emp_data.get("email", ""),
        "designation": emp_data.get("designation", ""),
        "function": emp_data.get("function", ""),
        "location": emp_data.get("location", ""),
        "bank_name": emp_data.get("bank_name", "N/A"),
        "bank_branch": emp_data.get("bank_branch", ""),
        "bank_number": emp_data.get("bank_number", "N/A"),
        "date_joined": emp_data.get("date_joined", ""),
        "ssnit_number": emp_data.get("ssnit_number", ""),
        "basic_salary": float(payroll.basic_salary or 0),
        "meals_monthly": float(payroll.meals_monthly or 0),
        "responsibility_allowance": float(payroll.responsibility_allowance or 0),
        "cola": float(payroll.cola or 0),
        "leave_allowance": float(payroll.leave_allowance or 0),
        "other_earnings": float(payroll.other_earnings or 0),
        "total_earnings": float(payroll.total_earnings or 0),
        "rent_monthly": float(payroll.rent_monthly or 0),
        "utility_monthly": float(payroll.utility_monthly or 0),
        "transport_monthly": float(payroll.transport_monthly or 0),
        "paye": float(payroll.paye or 0),
        "tithe": float(payroll.tithe or 0),
        "future_savings": float(payroll.future_savings or 0),
        "other_deductions": float(payroll.other_deductions or 0),
        "ssnit_deduction": float(payroll.ssnit_deduction or 0),
        "employee_pf": float(payroll.employee_pf or 0),
        "total_deductions": float(payroll.total_deductions or 0),
        "net_salary": float(payroll.net_salary or 0),
        "employer_contribution": float(payroll.employer_contribution or 0),
        "staff_category": staff_category,
        "pdf_generated": bool(payroll.pdf_generated) if hasattr(payroll, 'pdf_generated') else False,
        "loans": loans_data,
        "net_salary_words": net_words
    }

@app.get("/payslips/{payroll_id}/download")
async def download_payslip(payroll_id: int, request: Request, db: Session = Depends(get_db)):
    from app.models.payroll import PayrollRecord
    from app.services.pdf_service import generate_payslip_pdf, get_pdf_temp_dir
    from fastapi.responses import FileResponse
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Authenticate user
    try:
        current_user = get_current_user_web(request, db)
    except Exception:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    payroll = db.query(PayrollRecord).filter(PayrollRecord.id == payroll_id).first()
    if not payroll:
        raise HTTPException(status_code=404, detail="Payslip not found")
    
    logger.info(f"PAYSLIP DOWNLOAD: id={payroll_id}, employee={payroll.employee_name}, month={payroll.month}")
    
    # Generate PDF on demand if not already generated (or regenerated from temp)
    if not payroll.pdf_generated or not isinstance(payroll.pdf_generated, str) or not os.path.exists(payroll.pdf_generated):
        pdf_path = generate_payslip_pdf(db, payroll.id)
        if not pdf_path:
            raise HTTPException(status_code=500, detail="Failed to generate payslip PDF")
        payroll.pdf_generated = pdf_path
        db.commit()
    else:
        # If the file still exists and was generated in temp dir, serve it
        # Otherwise regenerate
        if not os.path.exists(payroll.pdf_generated):
            pdf_path = generate_payslip_pdf(db, payroll.id)
            if not pdf_path:
                raise HTTPException(status_code=500, detail="Failed to generate payslip PDF")
            payroll.pdf_generated = pdf_path
            db.commit()
    
    if not isinstance(payroll.pdf_generated, str) or not os.path.exists(payroll.pdf_generated):
        raise HTTPException(status_code=404, detail="PDF file not found")
    
    return FileResponse(payroll.pdf_generated, media_type="application/pdf", filename=os.path.basename(payroll.pdf_generated))

@app.post("/payslips/{payroll_id}/send-email")
async def send_payslip_email(payroll_id: int, request: Request, db: Session = Depends(get_db)):
    """Send payslip to employee via email"""
    try:
        current_user = get_current_user_web(request, db)
    except:
        # Return a simple HTML response for failed auth
        return await send_payslip_redirect(request, "Not authenticated. Please login again.", "error")
    
    from app.models.payroll import PayrollRecord
    from app.models.employee import Employee
    from app.services.email_service import EmailService
    from app.services.pdf_service import generate_payslip_pdf
    
    try:
        payroll = db.query(PayrollRecord).filter(PayrollRecord.id == payroll_id).first()
        if not payroll:
            return await send_payslip_redirect(request, "Payslip not found.", "error")
        
        employee = db.query(Employee).filter(Employee.name == payroll.employee_name).first()
        if not employee or not employee.email:
            return await send_payslip_redirect(request, "Employee email not configured.", "error")
        
        # Ensure PDF exists before sending
        if not payroll.pdf_generated or not isinstance(payroll.pdf_generated, str) or not os.path.exists(str(payroll.pdf_generated or '')):
            pdf_path = generate_payslip_pdf(db, payroll.id)
            if pdf_path:
                payroll.pdf_generated = pdf_path
                db.commit()
        
        if not payroll.pdf_generated or not isinstance(payroll.pdf_generated, str) or not os.path.exists(str(payroll.pdf_generated or '')):
            return await send_payslip_redirect(request, "Failed to generate PDF payslip.", "error")
        
        success, message = await EmailService.send_payslip(
            recipient_email=employee.email,
            employee_name=employee.name,
            month=payroll.month,
            pdf_path=payroll.pdf_generated,
            net_salary=payroll.net_salary
        )
        
        if success:
            return await send_payslip_redirect(request, f"Payslip sent to {employee.email}", "success")
        else:
            return await send_payslip_redirect(request, f"Failed to send email: {message}", "error")
    
    except Exception as e:
        return await send_payslip_redirect(request, f"Error sending payslip: {str(e)}", "error")


async def send_payslip_redirect(request: Request, msg: str, msg_type: str = "error"):
    """Redirect back to the send payslips page with a message"""
    referer = request.headers.get("referer", "/payslips/send")
    separator = "&" if "?" in referer else "?"
    return RedirectResponse(url=f"{referer}{separator}email_status={msg_type}&email_message={urllib.parse.quote(msg)}", status_code=303)

@app.get("/payslips/send", response_class=HTMLResponse)
@app.get("/payslips/send-all", response_class=HTMLResponse)
async def send_all_payslips_page(request: Request, db: Session = Depends(get_db), month: str = None, filter_by: str = "all", email_filter: str = None):
    """Page to send all payslips for a specific month. Auto-selects latest month."""
    try:
        current_user = get_current_user_web(request, db)
        from app.core.config import settings
        from app.models.payroll import PayrollRecord
        from app.models.employee import Employee
        from app.models.email_log import EmailLog
        from datetime import datetime as dt, timedelta
        
        # Get list of months with payslips (ordered newest first)
        payroll_months = db.query(PayrollRecord.month).distinct().order_by(PayrollRecord.month.desc()).all()
        months = [m[0] for m in payroll_months if m[0]]
        
        # Auto-detect latest month if none selected
        if not month and months:
            month = months[0]
        
        payslips = []
        if month:
            payslips = db.query(PayrollRecord).filter(PayrollRecord.month == month).all()
            
            # Attach employee info to each payslip using robust name matching
            all_employees = {e.name: e for e in db.query(Employee).all()}
            for p in payslips:
                emp_name = (p.employee_name or '').strip()
                # Try exact match first
                emp = all_employees.get(emp_name)
                if not emp:
                    # Try case-insensitive match
                    emp_name_upper = emp_name.upper()
                    for e_name, e_obj in all_employees.items():
                        if e_name.upper() == emp_name_upper:
                            emp = e_obj
                            break
                if not emp:
                    # Try normalized whitespace match
                    emp_name_norm = ' '.join(emp_name.split())
                    for e_name, e_obj in all_employees.items():
                        if ' '.join((e_name or '').split()) == emp_name_norm:
                            emp = e_obj
                            break
                p.employee = emp
            
            # Apply filter (support both filter_by and email_filter query params)
            effective_filter = email_filter if email_filter else filter_by
            if effective_filter == "has_email" or effective_filter == "with_email":
                payslips = [p for p in payslips if p.employee and p.employee.email and str(p.employee.email).strip()]
            elif effective_filter == "no_email":
                payslips = [p for p in payslips if not p.employee or not p.employee.email or not str(p.employee.email).strip()]
        
        # Get recent email logs
        email_logs = []
        try:
            email_logs = db.query(EmailLog).order_by(EmailLog.sent_at.desc()).limit(50).all()
        except Exception:
            email_logs = []
        
        template = templates.get_template("send_payslips.html")
        rendered = template.render({
            "user": current_user,
            "request": request,
            "months": months,
            "selected_month": month,
            "filter_by": filter_by,
        "email_filter": email_filter if email_filter else filter_by,
            "payslips": payslips,
            "email_logs": email_logs,
            "smtp_configured": bool(settings.smtp_server and settings.smtp_username),
            "smtp_server": settings.smtp_server or "",
            "smtp_port": settings.smtp_port or "",
            "smtp_username": settings.smtp_username or "",
            "smtp_password": settings.smtp_password or ""
        })
        return HTMLResponse(content=rendered, media_type="text/html")
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)

@app.post("/payslips/send-all")
async def send_all_payslips(request: Request, month: str = Form(...), db: Session = Depends(get_db)):
    """Send all payslips for a specific month"""
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.payroll import PayrollRecord
    from app.services.email_service import EmailService
    from app.services.pdf_service import generate_payslip_pdf
    
    try:
        # Get all payslips for the month
        payslips = db.query(PayrollRecord).filter(PayrollRecord.month == month).all()
        
        if not payslips:
            template = templates.get_template("send_payslips.html")
            rendered = template.render({"user": current_user, "error": f"No payslips found for {month}"})
            return HTMLResponse(content=rendered, media_type="text/html")
        
        # First generate PDFs for any records that don't have them
        for p in payslips:
            if not p.pdf_generated or not os.path.exists(p.pdf_generated):
                pdf_path = generate_payslip_pdf(db, p.id)
                if pdf_path:
                    p.pdf_generated = pdf_path
        db.commit()
        
        # Send all payslips
        results = await EmailService.send_bulk_payslips(payslips, db)
        
        template = templates.get_template("send_payslips.html")
        rendered = template.render({
            "user": current_user, 
            "success": f"Sent {results['successful']} payslips. Failed: {results['failed']}",
            "results": results
        })
        return HTMLResponse(content=rendered, media_type="text/html")
    
    except Exception as e:
        template = templates.get_template("send_payslips.html")
        rendered = template.render({"user": current_user, "error": f"Error sending payslips: {str(e)}"})
        return HTMLResponse(content=rendered, media_type="text/html")

@app.post("/payslips/generate-all", response_class=JSONResponse)
async def generate_all_payslips(request: Request, month: str = Form(""), db: Session = Depends(get_db)):
    """Generate PDFs for all payslips in a given month, reusing existing files where possible."""
    from app.models.payroll import PayrollRecord
    from app.services.pdf_service import generate_payslip_pdf
    import os
    
    if month and month != "__ALL__":
        records = db.query(PayrollRecord).filter(PayrollRecord.month == month).all()
    else:
        records = db.query(PayrollRecord).all()
        month = "All Months"
    if not records:
        return JSONResponse(content={"success": False, "message": f"No payslips found for {month}"})
    
    generated = 0
    reused = 0
    failed = 0
    for r in records:
        try:
            # Reuse existing PDF if it exists and is valid
            existing_path = r.pdf_generated if hasattr(r, 'pdf_generated') else None
            if existing_path and os.path.exists(existing_path):
                reused += 1
                continue
            
            pdf_path = generate_payslip_pdf(db, r.id)
            if pdf_path:
                r.pdf_generated = pdf_path
                generated += 1
            else:
                failed += 1
        except:
            failed += 1
    db.commit()
    
    message = f"Generated {generated} new PDF(s), reused {reused} existing. Failed: {failed}."
    return JSONResponse(content={"success": True, "message": message})


@app.post("/payslips/{payroll_id}/generate-pdf", response_class=JSONResponse)
@app.post("/payslips/generate-single/{payroll_id}", response_class=JSONResponse)
async def generate_single_payslip(payroll_id: int, db: Session = Depends(get_db)):
    """Generate PDF for a single payslip. Reuses existing file if available."""
    from app.models.payroll import PayrollRecord
    from app.services.pdf_service import generate_payslip_pdf
    import os
    
    r = db.query(PayrollRecord).filter(PayrollRecord.id == payroll_id).first()
    if not r:
        return JSONResponse(content={"success": False, "message": "Payslip not found"}, status_code=404)
    
    # Check if existing PDF is still valid
    existing_path = r.pdf_generated if hasattr(r, 'pdf_generated') else None
    if existing_path and os.path.exists(existing_path):
        return JSONResponse(content={"success": True, "message": "PDF already exists", "path": existing_path})
    
    try:
        pdf_path = generate_payslip_pdf(db, r.id)
        if pdf_path:
            r.pdf_generated = pdf_path
            db.commit()
            return JSONResponse(content={"success": True, "message": "PDF generated successfully", "path": pdf_path})
        else:
            return JSONResponse(content={"success": False, "message": "PDF generation failed"}, status_code=500)
    except Exception as e:
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)




# === LOAN MANAGEMENT ===
@app.get("/loans", response_class=HTMLResponse)
async def loans_page(request: Request, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.loan import Loan
    from app.models.employee import Employee
    
    loans = db.query(Loan).order_by(Loan.id.desc()).all()
    all_employees = db.query(Employee).filter(Employee.name.isnot(None)).all()
    
    template = templates.get_template("loans.html")
    rendered = template.render({
        "user": current_user,
        "loans": loans,
        "all_employees": all_employees
    })
    return HTMLResponse(content=rendered, media_type="text/html")


@app.get("/api/employee-names")
async def loans_page(request: Request, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.loan import Loan
    from app.models.employee import Employee
    from app.services.cache_service import get_cache, set_cache
    import json
    
    cache_key = "loans_data"
    cached = get_cache(cache_key)
    if cached:
        loans, employees_list, active_count, total_outstanding_sum, total_loaned_sum, completed_count = cached
        employees_json = json.dumps(employees_list)
    else:
        loans = db.query(Loan).order_by(Loan.created_at.desc()).all()
        employees = db.query(Employee).all()
        
        # Prepare employees JSON for autocomplete
        employees_list = [{"name": e.name, "employee_number": e.employee_number or ""} for e in employees if e.name]
        employees_json = json.dumps(employees_list)
        
        # Stats
        active_count = db.query(Loan).filter(Loan.status == "Active").count()
        total_outstanding = db.query(Loan).filter(Loan.status == "Active").with_entities(Loan.balance).all()
        total_outstanding_sum = sum(b[0] or 0 for b in total_outstanding)
        total_loaned = db.query(Loan).with_entities(Loan.loan_amount).all()
        total_loaned_sum = sum(l[0] or 0 for l in total_loaned)
        completed_count = db.query(Loan).filter(Loan.status == "Completed").count()
        set_cache(cache_key, (loans, employees_list, active_count, total_outstanding_sum, total_loaned_sum, completed_count), ttl_seconds=120)
    
    stats = {
        "active_count": active_count,
        "total_outstanding": total_outstanding_sum,
        "total_loaned": total_loaned_sum,
        "completed_count": completed_count
    }
    
    import json as _json_mod
    loans_json = _json_mod.dumps([{
        "id": l.id,
        "employee_name": l.employee_name,
        "bank_name": l.bank_name,
        "loan_amount": l.loan_amount,
        "interest_amount": l.interest_amount,
        "total_receivable": l.total_receivable,
        "monthly_deduction": l.monthly_deduction,
        "amount_paid": l.amount_paid,
        "months_to_pay": l.months_to_pay,
        "months_paid": l.months_paid,
        "balance": l.balance,
        "status": l.status,
        "notes": l.notes
    } for l in loans])
    
    template = templates.get_template("loans.html")
    rendered = template.render({
        "user": current_user,
        "loans": loans,
        "stats": stats,
        "employees_json": employees_json,
        "loans_json": loans_json,
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error")
    })
    return HTMLResponse(content=rendered, media_type="text/html")

@app.post("/loans/bulk-add")
async def bulk_add_loans(request: Request, db: Session = Depends(get_db)):
    """Register loans for multiple employees with individual amounts."""
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.loan import Loan
    
    # Parse form data manually since we have paired arrays
    form = await request.form()
    employee_names = form.getlist("employee_names")
    loan_amounts = form.getlist("loan_amounts")
    bank_name = form.get("bank_name", "") or ""
    months_to_pay = int(form.get("months_to_pay", 12))
    interest_amount = float(form.get("interest_amount", 0)) if form.get("interest_amount") else 0
    
    added = 0
    for i, name in enumerate(employee_names):
        if not name.strip():
            continue
        
        # Get individual loan amount for this employee
        loan_amount = float(loan_amounts[i]) if i < len(loan_amounts) else 0
        if loan_amount <= 0:
            continue
        
        total_receivable = loan_amount + interest_amount
        monthly_deduction = total_receivable / months_to_pay if months_to_pay > 0 else total_receivable
        
        loan = Loan(
            employee_name=name.strip(),
            bank_name=bank_name,
            loan_amount=loan_amount,
            months_to_pay=months_to_pay,
            interest_amount=interest_amount,
            total_receivable=total_receivable,
            monthly_deduction=monthly_deduction,
            amount_paid=0,
            months_paid=0,
            balance=total_receivable,
            status="Active"
        )
        db.add(loan)
        added += 1
    
    db.commit()
    return RedirectResponse(url=f"/loans?success=Bulk+loan+registered+for+{added}+employee(s)", status_code=303)

@app.post("/loans/add")
async def add_loan(request: Request, employee_name: str = Form(...), bank_name: str = Form(default=""),
                   loan_amount: float = Form(default=0), interest_amount: float = Form(default=0),
                   months_to_pay: int = Form(default=1), notes: str = Form(default=""),
                   db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.loan import Loan
    
    try:
        total_receivable = loan_amount + interest_amount
        monthly_deduction = total_receivable / months_to_pay if months_to_pay > 0 else total_receivable
        
        loan = Loan(
            employee_name=employee_name,
            bank_name=bank_name,
            loan_amount=loan_amount,
            interest_amount=interest_amount,
            months_to_pay=months_to_pay,
            total_receivable=total_receivable,
            monthly_deduction=monthly_deduction,
            amount_paid=0,
            months_paid=0,
            balance=total_receivable,
            status="Active",
            notes=notes
        )
        db.add(loan)
        db.commit()
        
        return RedirectResponse(url="/loans?success=Loan+recorded+successfully", status_code=303)
    except Exception as e:
        db.rollback()
        return RedirectResponse(url=f"/loans?error={str(e)}", status_code=303)

@app.get("/api/employee-names")
async def get_employee_names(request: Request, db: Session = Depends(get_db)):
    """Return list of employee names for autocomplete"""
    try:
        current_user = get_current_user_web(request, db)
    except Exception:
        return JSONResponse(content={"names": []})
    
    from app.models.employee import Employee
    employees = db.query(Employee).filter(Employee.name.isnot(None)).all()
    names = [e.name for e in employees if e.name]
    return JSONResponse(content={"names": names})

@app.post("/loans/pay")
async def pay_loan(request: Request, loan_id: int = Form(...), db: Session = Depends(get_db)):
    """Record one monthly payment. Reduces months remaining and increases amount paid."""
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.loan import Loan
    
    loan = db.query(Loan).filter(Loan.id == loan_id).first()
    if not loan:
        return RedirectResponse(url="/loans?error=Loan+not+found", status_code=303)
    
    try:
        payment_amount = loan.monthly_deduction or 0
        loan.amount_paid = (loan.amount_paid or 0) + payment_amount
        loan.months_paid = (loan.months_paid or 0) + 1
        loan.balance = max(0, (loan.total_receivable or 0) - (loan.amount_paid or 0))
        
        if loan.balance <= 0:
            loan.status = "Completed"
        
        db.commit()
        return RedirectResponse(url=f"/loans?success=Payment+recorded+for+{loan.employee_name}", status_code=303)
    except Exception as e:
        db.rollback()
        return RedirectResponse(url=f"/loans?error={str(e)}", status_code=303)

@app.post("/loans/complete/{loan_id}")
async def complete_loan(request: Request, loan_id: int, db: Session = Depends(get_db)):
    """Mark a loan as completed."""
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.loan import Loan
    
    loan = db.query(Loan).filter(Loan.id == loan_id).first()
    if not loan:
        return RedirectResponse(url="/loans?error=Loan+not+found", status_code=303)
    
    loan.status = "Completed"
    loan.balance = 0
    loan.amount_paid = loan.total_receivable or 0
    if loan.months_to_pay:
        loan.months_paid = loan.months_to_pay
    db.commit()
    return RedirectResponse(url=f"/loans?success=Loan+{loan.employee_name}+marked+as+completed", status_code=303)

@app.post("/loans/{loan_id}/default")
async def default_loan(loan_id: int, db: Session = Depends(get_db)):
    from app.models.loan import Loan
    
    loan = db.query(Loan).filter(Loan.id == loan_id).first()
    if not loan:
        return JSONResponse(content={"success": False, "message": "Loan not found"}, status_code=404)
    
    loan.status = "Defaulted"
    db.commit()
    return JSONResponse(content={"success": True, "message": "Loan marked as defaulted"})

@app.post("/loans/delete/{loan_id}")
async def delete_loan(request: Request, loan_id: int, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.loan import Loan
    
    loan = db.query(Loan).filter(Loan.id == loan_id).first()
    if loan:
        db.delete(loan)
        db.commit()
        return RedirectResponse(url="/loans?success=Loan+deleted", status_code=303)
    return RedirectResponse(url="/loans?error=Loan+not+found", status_code=303)


# === EMAIL SETTINGS ===
@app.get("/settings/email", response_class=HTMLResponse)
async def email_settings_page(request: Request, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.core.config import settings
    
    # Determine security type - default to tls for port 587
    port = settings.smtp_port or 587
    if port == 587:
        smtp_security = "tls"
    elif port == 465:
        smtp_security = "ssl"
    else:
        smtp_security = "none"
    
    template = templates.get_template("email_settings.html")
    rendered = template.render({
        "user": current_user,
        "message": request.query_params.get("success"),
        "error": request.query_params.get("error"),
        "smtp_configured": bool(settings.smtp_server and settings.smtp_username),
        "smtp_server": settings.smtp_server or "",
        "smtp_port": port,
        "smtp_username": settings.smtp_username or "",
        "smtp_password": settings.smtp_password or "",
        "smtp_security": smtp_security
    })
    return HTMLResponse(content=rendered, media_type="text/html")

@app.post("/settings/email")
async def email_settings_save(request: Request, smtp_server: str = Form(...), smtp_port: int = Form(...),
                              smtp_username: str = Form(...), smtp_password: str = Form(...),
                              smtp_security: str = Form(...), db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
        # Save to .env file
    env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.env')
    try:
        with open(env_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        updates = {
            'smtp_server': smtp_server,
            'smtp_port': str(smtp_port),
            'smtp_username': smtp_username,
            'smtp_password': smtp_password,
            'smtp_security': smtp_security
        }
        
        updated_keys = set()
        new_lines = []
        for line in lines:
            stripped = line.strip()
            if '=' in stripped:
                key = stripped.split('=')[0].strip().lower()
                if key in updates:
                    new_lines.append(f'{key}={updates[key]}\n')
                    updated_keys.add(key)
                else:
                    new_lines.append(line)
            else:
                new_lines.append(line)
        
        # Add any missing keys
        for key, val in updates.items():
            if key not in updated_keys:
                new_lines.append(f'{key}={val}\n')
        
        with open(env_path, 'w', encoding='utf-8') as f:
            f.writelines(new_lines)
        
        return RedirectResponse(url="/settings/email?success=Settings+saved+to+.env+file.+Restart+server+to+apply.", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/settings/email?error=Failed+to+save:+{str(e)}", status_code=303)

@app.post("/settings/email/test")
async def test_email_connection(request: Request, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    from app.core.config import settings
    
    
    if not settings.smtp_server or not settings.smtp_username:
        raise HTTPException(status_code=400, detail="SMTP not configured. Check your .env file.")
    
    return {"status": "ok", "message": f"SMTP configured: {settings.smtp_server}:{settings.smtp_port}"}


# === PAYROLL REPORTS ===
@app.get("/reports", response_class=HTMLResponse)
async def reports_page(request: Request, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    from app.models.payroll import PayrollRecord
    from app.models.employee import Employee
    from app.services.cache_service import get_cache, set_cache
        
    cache_key = "reports_data"
    cached = get_cache(cache_key)
    if cached:
        stats, monthly_data = cached
    else:
        total_records = db.query(PayrollRecord).count()
        total_employees = db.query(Employee).count()
        
        total_payroll_data = db.query(PayrollRecord.net_salary).all()
        total_payroll = sum(p[0] or 0 for p in total_payroll_data)
        avg_net = total_payroll / total_records if total_records > 0 else 0
        
        stats = {
            "total_records": total_records,
            "total_employees": total_employees,
            "total_payroll": total_payroll,
            "avg_net": avg_net
        }
        
        # Monthly breakdown
        months = db.query(PayrollRecord.month).distinct().order_by(PayrollRecord.month.desc()).all()
        monthly_data = []
        for m in months:
            if m[0]:
                records = db.query(PayrollRecord).filter(PayrollRecord.month == m[0]).all()
                monthly_data.append({
                    "month": m[0],
                    "count": len(records),
                    "total": sum(r.net_salary or 0 for r in records)
                })
        set_cache(cache_key, (stats, monthly_data), ttl_seconds=120)
    
    # Build month list for template
    months_list = [m["month"] for m in monthly_data]
    
    template = templates.get_template("reports.html")
    rendered = template.render({
        "user": current_user,
        "stats": stats,
        "employees_count": stats["total_employees"],
        "months": months_list,
        "total_payroll": round(stats["total_payroll"], 2),
        "payslip_count": stats["total_records"],
        "monthly_data": monthly_data,
        "error": None
    })
    return HTMLResponse(content=rendered, media_type="text/html")


# === REPORT GENERATION ===
@app.post("/reports/generate")
async def generate_report(request: Request, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    form = await request.form()
    report_type = form.get("type", "payroll_summary")
    month = form.get("month", "")
    fmt = form.get("format", "excel")
    
    from app.models.payroll import PayrollRecord
    from app.models.employee import Employee
    from app.services.cache_service import get_cache, set_cache
    
    # Query payroll records
    query = db.query(PayrollRecord)
    if month:
        query = query.filter(PayrollRecord.month == month)
    records = query.order_by(PayrollRecord.employee_name).all()
    
    if fmt == "excel":
        import pandas as pd
        import io
        from openpyxl import Workbook
        
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.replace("_", " ").title()
        
        if report_type == "payroll_summary":
            headers = ["Employee", "Month", "Category", "Basic Salary", "Meals", "Responsibility Allowance", "COLA",
                       "Leave Allowance", "Total Earnings", "SSNIT 5.5%", "PAYE", "Tithe", "Future Savings",
                       "Employee PF", "Other Deductions", "Total Deductions", "Net Salary"]
            ws.append(headers)
            for r in records:
                ws.append([r.employee_name, r.month, r.staff_category or "pastoral", r.basic_salary, r.meals_monthly, r.responsibility_allowance,
                          r.cola, r.leave_allowance, r.total_earnings, r.ssnit_deduction, r.paye, r.tithe, r.future_savings,
                          r.employee_pf, r.other_deductions, r.total_deductions, r.net_salary])
        elif report_type == "employee_list":
            employees = db.query(Employee).order_by(Employee.name).all()
            headers = ["Name", "Email", "Employee Number", "Designation", "Location", "Bank", "Bank Number"]
            ws.append(headers)
            for e in employees:
                ws.append([e.name, e.email, e.employee_number, e.designation or e.function, e.location, e.bank_name, e.bank_number])
        elif report_type == "deductions":
            headers = ["Employee", "Month", "Category", "SSNIT 5.5%", "PAYE", "Tithe", "Future Savings", "Employee PF", "Other Deductions", "Total Deductions"]
            ws.append(headers)
            for r in records:
                ws.append([r.employee_name, r.month, r.staff_category or "pastoral", r.ssnit_deduction, r.paye, r.tithe, r.future_savings, r.employee_pf, r.other_deductions, r.total_deductions])
        elif report_type == "bank_payments":
            headers = ["Employee", "Bank", "Bank Number", "Net Salary", "Month"]
            ws.append(headers)
            for r in records:
                emp = db.query(Employee).filter(Employee.name == r.employee_name).first()
                ws.append([r.employee_name, emp.bank_name if emp else "", emp.bank_number if emp else "", r.net_salary, r.month])
        else:
            headers = ["Employee", "Month", "Basic Salary", "Total Earnings", "Total Deductions", "Net Salary"]
            ws.append(headers)
            for r in records:
                ws.append([r.employee_name, r.month, r.basic_salary, r.total_earnings, r.total_deductions, r.net_salary])
        
        wb.save(output)
        output.seek(0)
        filename = f"{report_type}_{month or 'all'}.xlsx"
        return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename={filename}"})
    else:
        # CSV format
        import csv
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        
        if report_type == "payroll_summary":
            writer.writerow(["Employee", "Month", "Category", "Basic Salary", "Total Earnings", "SSNIT 5.5%", "Total Deductions", "Net Salary"])
            for r in records:
                writer.writerow([r.employee_name, r.month, r.staff_category or "pastoral", r.basic_salary, r.total_earnings, r.ssnit_deduction, r.total_deductions, r.net_salary])
        elif report_type == "employee_list":
            employees = db.query(Employee).order_by(Employee.name).all()
            writer.writerow(["Name", "Email", "Employee Number", "Designation"])
            for e in employees:
                writer.writerow([e.name, e.email, e.employee_number, e.designation])
        elif report_type == "deductions":
            writer.writerow(["Employee", "Month", "Category", "SSNIT 5.5%", "PAYE", "Tithe", "Future Savings", "Employee PF", "Other Deductions", "Total Deductions"])
            for r in records:
                writer.writerow([r.employee_name, r.month, r.staff_category or "pastoral", r.ssnit_deduction, r.paye, r.tithe, r.future_savings, r.employee_pf, r.other_deductions, r.total_deductions])
        elif report_type == "bank_payments":
            writer.writerow(["Employee", "Bank", "Bank Number", "Net Salary", "Month"])
            for r in records:
                emp = db.query(Employee).filter(Employee.name == r.employee_name).first()
                writer.writerow([r.employee_name, emp.bank_name if emp else "", emp.bank_number if emp else "", r.net_salary, r.month])
        else:
            writer.writerow(["Employee", "Month", "Basic Salary", "Total Earnings", "Total Deductions", "Net Salary"])
            for r in records:
                writer.writerow([r.employee_name, r.month, r.basic_salary, r.total_earnings, r.total_deductions, r.net_salary])
        
        output.seek(0)
        filename = f"{report_type}_{month or 'all'}.csv"
        return Response(content=output.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.get("/reports/month-summary")
async def month_summary(request: Request, month: str = "", db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    
    if not month:
        return JSONResponse({"error": "Month parameter required"}, status_code=400)
    
    from app.models.payroll import PayrollRecord
        
    try:
        records = db.query(PayrollRecord).filter(PayrollRecord.month == month).all()
        count = len(records)
        total_net = sum(r.net_salary or 0 for r in records)
        total_deductions = sum(r.total_deductions or 0 for r in records)
        total_earnings = sum(r.total_earnings or 0 for r in records)
        avg_net = round(total_net / count, 2) if count > 0 else 0
        
        return JSONResponse({
            "month": month,
            "count": count,
            "total_net": round(total_net, 2),
            "total_earnings": round(total_earnings, 2),
            "total_deductions": round(total_deductions, 2),
            "avg_net": avg_net
        })
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


# === UPLOAD HISTORY ===
@app.get("/reports/history", response_class=HTMLResponse)
async def upload_history_page(request: Request, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=303)
    
    try:
        from app.models.upload_history import UploadHistory
    except ImportError:
        template = templates.get_template("upload_history.html")
        rendered = template.render({
            "user": current_user,
            "uploads": [],
            "error": "Upload history module not available. Please run database migrations."
        })
        return HTMLResponse(content=rendered, media_type="text/html")
    
    from app.services.cache_service import get_cache, set_cache
    
    cache_key = "upload_hist"
    cached = get_cache(cache_key)
    if cached:
        uploads_list = cached
    else:
        try:
            uploads = db.query(UploadHistory).order_by(UploadHistory.id.desc()).limit(500).all()
        except Exception as e:
            template = templates.get_template("upload_history.html")
            rendered = template.render({
                "user": current_user,
                "uploads": [],
                "error": f"Database error: {str(e)[:100]}"
            })
            return HTMLResponse(content=rendered, media_type="text/html")
        
        # Convert to serializable format for template
        uploads_list = []
        for u in uploads:
            ts = u.timestamp
            
            uploads_list.append({
                "id": u.id,
                "file_name": u.file_name,
                "uploaded_by": u.uploaded_by,
                "month": u.month,
                "status": u.status,
                "timestamp": ts
            })
        
        set_cache(cache_key, uploads_list, ttl_seconds=120)
    
    template = templates.get_template("upload_history.html")
    rendered = template.render({
        "user": current_user,
        "uploads": uploads_list,
        "error": None
    })
    return HTMLResponse(content=rendered, media_type="text/html")




@app.get("/api/employees/{employee_id}/detail")
async def employee_detail(request: Request, employee_id: int, db: Session = Depends(get_db)):
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return {"error": "Not authenticated"}
    
    from app.models.employee import Employee
    
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        return JSONResponse(content={"error": "Employee not found"}, status_code=404)
    
    return {
        "id": emp.id,
        "employee_number": emp.employee_number,
        "name": emp.name,
        "email": emp.email,
        "function": emp.function,
        "designation": emp.designation,
        "location": emp.location,
        "date_joined": emp.date_joined.isoformat() if emp.date_joined else None,
        "ssnit_number": emp.ssnit_number,
        "tax_relief": emp.tax_relief,
        "employer_contribution": emp.employer_contribution or 0,
        "bank_name": emp.bank_name,
        "bank_number": emp.bank_number,
        "bank_branch": emp.bank_branch
    }

@app.post("/payslips/{payroll_id}/send-single")
async def send_single_payslip(request: Request, payroll_id: int, db: Session = Depends(get_db)):
    """Send a single payslip to an employee via email."""
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return JSONResponse(content={"success": False, "error": "Not authenticated"}, status_code=401)

    from app.models.payroll import PayrollRecord
    from app.models.employee import Employee
    from app.services.pdf_service import generate_payslip_pdf
    from app.services.email_service import send_payslip_email

    record = db.query(PayrollRecord).filter(PayrollRecord.id == payroll_id).first()
    if not record:
        return JSONResponse(content={"success": False, "error": "Payslip not found"}, status_code=404)

    emp = db.query(Employee).filter(Employee.name == record.employee_name).first()
    if not emp:
        return JSONResponse(content={"success": False, "error": "Employee not found"}, status_code=404)

    if not emp.email:
        return JSONResponse(content={"success": False, "error": "No email address for this employee"})

    # Generate PDF if not exists
    pdf_path = record.pdf_generated
    if not pdf_path or not os.path.exists(pdf_path):
        pdf_path = generate_payslip_pdf(db, record.id)
        if pdf_path:
            record.pdf_generated = pdf_path
            db.commit()

    if not pdf_path or not os.path.exists(pdf_path):
        return JSONResponse(content={"success": False, "error": "Failed to generate PDF"})

    # Send email
    try:
        success, error_msg = await send_payslip_email(emp.email, emp.name, pdf_path, record.month)
        if success:
            return JSONResponse(content={"success": True, "message": "Payslip sent to " + emp.email})
        else:
            print(f"Email send failed for {emp.email}: {error_msg}")
            return JSONResponse(content={"success": False, "error": error_msg})
    except Exception as e:
        import traceback
        print(f"send_single_payslip ERROR: {e}")
        traceback.print_exc()
        return JSONResponse(content={"success": False, "error": str(e)})


@app.post("/payslips/{log_id}/resend")
async def resend_payslip(request: Request, log_id: int, db: Session = Depends(get_db)):
    """Resend a failed payslip by EmailLog ID."""
    try:
        current_user = get_current_user_web(request, db)
    except HTTPException:
        return JSONResponse(content={"success": False, "error": "Not authenticated"}, status_code=401)

    from app.models.email_log import EmailLog
    from app.models.payroll import PayrollRecord
    from app.models.employee import Employee
    from app.services.pdf_service import generate_payslip_pdf
    from app.services.email_service import send_payslip_email
    from datetime import datetime

    log_entry = db.query(EmailLog).filter(EmailLog.id == log_id).first()
    if not log_entry:
        return JSONResponse(content={"success": False, "error": "Log entry not found"}, status_code=404)

    emp = db.query(Employee).filter(Employee.id == log_entry.employee_id).first()
    if not emp:
        return JSONResponse(content={"success": False, "error": "Employee not found"})

    record = db.query(PayrollRecord).filter(
        PayrollRecord.month == log_entry.month,
        PayrollRecord.employee_name == log_entry.employee_name
    ).first()
    if not record:
        return JSONResponse(content={"success": False, "error": "Payroll record not found"})

    # Generate PDF if not exists
    pdf_path = record.pdf_generated
    if not pdf_path or not os.path.exists(pdf_path):
        pdf_path = generate_payslip_pdf(db, record.id)
        if pdf_path:
            record.pdf_generated = pdf_path
            db.commit()

    if not pdf_path or not os.path.exists(pdf_path):
        return JSONResponse(content={"success": False, "error": "Failed to generate PDF"})

    # Send email
    success, error_msg = await send_payslip_email(emp.email, emp.name, pdf_path, record.month)

    # Update log entry
    log_entry.status = "sent" if success else "failed"
    log_entry.error_message = None if success else error_msg
    log_entry.sent_at = datetime.utcnow()
    db.commit()

    if success:
        return JSONResponse(content={"success": True, "message": "Payslip resent to " + emp.email})
    else:
        return JSONResponse(content={"success": False, "error": error_msg})


@app.get("/logout")
async def logout(request: Request):
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie(key="access_token")
    return response

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=True)
